import cv2
import pytesseract
import pandas as pd
import numpy as np
import re
import os
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 중요 필드 정의 - 실제 바잉 AMD가 주로 확인하는 항목들
KEY_FIELDS = {
    'payment_terms': ['payment terms', '결제 조건', 'payment method'],
    'incoterms': ['incoterms', 'incoterm'],
    'total_price': ['total', 'grand total', '총액', '합계'],
    'reference': ['your reference', '주문 참조', 'reference'],
    'date': ['date', '날짜'],
    'season': ['season', '시즌'],
    'brand': ['brand', '브랜드'],
    'product_line': ['product line', '제품 라인']
}

# 이미지 전처리 개선 함수
def preprocess_image_for_ocr(image_path, enhance_key_areas=True):
    """
    OCR 성능 향상을 위한 이미지 전처리 함수
    
    Parameters:
    -----------
    image_path : str
        이미지 파일 경로
    enhance_key_areas : bool
        중요 영역 강화 여부
        
    Returns:
    --------
    numpy.ndarray
        전처리된 이미지
    """
    # 이미지 로드
    image = cv2.imread(image_path)
    if image is None:
        raise ValueError(f"이미지를 로드할 수 없습니다: {image_path}")
    
    # 그레이스케일 변환
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # 노이즈 제거 및 대비 향상
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    
    # CLAHE(Contrast Limited Adaptive Histogram Equalization) 적용
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(blur)
    
    # 적응형 이진화 적용
    thresh = cv2.adaptiveThreshold(
        enhanced, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
        cv2.THRESH_BINARY, 11, 2
    )
    
    # 중요 영역 강화 (표와 텍스트 경계 선명화)
    if enhance_key_areas:
        # 형태학적 연산으로 텍스트 강화
        kernel = np.ones((1, 1), np.uint8)
        thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
        
        # 엣지 검출로 표 구조 강화
        edges = cv2.Canny(thresh, 50, 150)
        dilated_edges = cv2.dilate(edges, kernel, iterations=1)
        
        # 원본 이진화 이미지와 엣지 결합
        result = cv2.bitwise_and(thresh, thresh, mask=~dilated_edges)
    else:
        result = thresh
    
    return result

# OCR 실행 함수 개선
def extract_text_with_enhanced_ocr(image_path, lang='eng+kor'):
    """
    개선된 OCR 텍스트 추출 함수
    
    Parameters:
    -----------
    image_path : str
        이미지 파일 경로
    lang : str
        OCR 언어 설정
        
    Returns:
    --------
    str
        추출된 텍스트
    """
    # 이미지 전처리
    processed_image = preprocess_image_for_ocr(image_path)
    
    # 기본 OCR 설정
    basic_config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
    text = pytesseract.image_to_string(processed_image, lang=lang, config=basic_config)
    
    # 표 감지 최적화 OCR 설정
    table_config = r'--oem 3 --psm 4 -c preserve_interword_spaces=1'
    table_text = pytesseract.image_to_string(processed_image, lang=lang, config=table_config)
    
    # 텍스트 행 수 비교 후 더 많은 행을 가진 결과 선택
    if table_text.count('\n') > text.count('\n'):
        text = table_text
    
    # 중요 필드 추출 최적화를 위한 OCR 재시도
    # 숫자와 기호에 최적화된 설정
    number_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist="0123456789.,-:/%€$£¥ABCDEFGHIJKLMNOPQRSTUVWXYZ abcdefghijklmnopqrstuvwxyz"'
    number_text = pytesseract.image_to_string(processed_image, lang=lang, config=number_config)
    
    # 두 결과 결합
    combined_text = text + "\n" + number_text
    
    # 중복 제거
    lines = combined_text.split('\n')
    unique_lines = []
    seen = set()
    
    for line in lines:
        line = line.strip()
        if line and line not in seen:
            seen.add(line)
            unique_lines.append(line)
    
    return '\n'.join(unique_lines)

# 중요 정보 추출 함수
def extract_key_information(text):
    """
    텍스트에서 중요 정보 추출
    
    Parameters:
    -----------
    text : str
        OCR로 추출한 텍스트
        
    Returns:
    --------
    dict
        추출된 중요 정보
    """
    key_info = {
        'payment_terms': '',
        'incoterms': '',
        'total_price': '',
        'reference': '',
        'date': '',
        'season': '',
        'brand': '',
        'product_line': '',
        'items': []
    }
    
    lines = text.split('\n')
    
    # 각 줄 검사
    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
        
        # 각 중요 필드 검색
        for field, keywords in KEY_FIELDS.items():
            for keyword in keywords:
                pattern = re.compile(f"{keyword}\\s*:?\\s*(.+)", re.IGNORECASE)
                match = pattern.search(line)
                if match:
                    key_info[field] = match.group(1).strip()
                    break
        
        # 특정 형식 필드 처리
        # Payment Terms
        if re.search(r'payment\s+terms', line, re.IGNORECASE):
            if i+1 < len(lines) and lines[i+1].strip():
                payment_terms_match = re.search(r'(\d+%.*\d+%.*days)', lines[i+1], re.IGNORECASE)
                if payment_terms_match:
                    key_info['payment_terms'] = payment_terms_match.group(1)
        
        # 가격 정보 (EUR 값)
        eur_match = re.search(r'EUR\s*([\d,]+\.\d{2})', line, re.IGNORECASE)
        if eur_match and ('total' in line.lower() or 'grand' in line.lower()):
            key_info['total_price'] = eur_match.group(0)
        
        # 날짜 형식 (xx-xx-xxxx)
        date_match = re.search(r'\b(\d{2}[-/]\d{2}[-/]\d{4})\b', line)
        if date_match and not key_info['date']:
            key_info['date'] = date_match.group(1)
        
        # 제품 항목 추출
        # 스타일 코드 패턴: FTVRM 또는 #FTVRM으로 시작
        style_match = re.search(r'(?:#)?FTVRM\w{1,10}(?:\d{5})?', line, re.IGNORECASE)
        model_match = re.search(r'(AJ\d{3,4})\s*-\s*([A-Za-z0-9 ]+)', line, re.IGNORECASE)
        
        if style_match or model_match:
            style_code = style_match.group(0) if style_match else ""
            model = model_match.group(0) if model_match else ""
            
            # 가격 정보 검색
            price_match = re.search(r'EUR\s*(\d+\.\d{2})', line)
            price = price_match.group(1) if price_match else ""
            
            # 다음 줄에서 사이즈 및 수량 정보 추출 시도
            sizes = []
            quantities = []
            
            if i+2 < len(lines):
                # 사이즈 행
                size_line = lines[i+1].strip()
                size_matches = re.findall(r'\b(39|40|41|42|43|44|45|46)\b', size_line)
                if size_matches:
                    sizes = size_matches
                    
                    # 수량 행
                    qty_line = lines[i+2].strip()
                    qty_matches = re.findall(r'\b\d+\b', qty_line)
                    if qty_matches and len(qty_matches) >= len(sizes):
                        quantities = qty_matches[:len(sizes)]
            
            if style_code or model:
                item = {
                    'style_code': style_code,
                    'model': model,
                    'price': price,
                    'sizes': sizes,
                    'quantities': quantities
                }
                key_info['items'].append(item)
    
    # 브랜드 명 추출 (TOGA VIRILIS)
    brand_match = re.search(r'TOGA\s+VIRILIS', text, re.IGNORECASE)
    if brand_match:
        key_info['brand'] = brand_match.group(0)
    
    # 시즌 정보 추출 (예: 2024SS, SS24 등)
    season_match = re.search(r'\b(20\d{2}SS|SS\d{2})\b', text)
    if season_match:
        key_info['season'] = season_match.group(0)
    
    return key_info

# PDF에서 중요 정보 추출
def extract_info_from_pdf(pdf_path, output_text_path=None):
    """
    PDF 파일에서 중요 정보 추출
    
    Parameters:
    -----------
    pdf_path : str
        PDF 파일 경로
    output_text_path : str, optional
        추출된 텍스트 저장 경로
        
    Returns:
    --------
    dict
        추출된 중요 정보
    """
    # PDF를 이미지로 변환
    try:
        images = convert_from_path(pdf_path, dpi=300)
        all_text = ""
        
        for i, image in enumerate(images):
            image_path = f"temp_page_{i}.jpg"
            image.save(image_path, "JPEG", quality=100)
            
            # OCR 실행
            page_text = extract_text_with_enhanced_ocr(image_path)
            all_text += page_text + "\n\n--- Page Break ---\n\n"
            
            # 임시 이미지 파일 삭제
            if os.path.exists(image_path):
                os.remove(image_path)
        
        # 텍스트 파일로 저장 (선택 사항)
        if output_text_path:
            with open(output_text_path, 'w', encoding='utf-8') as f:
                f.write(all_text)
        
        # 중요 정보 추출
        key_info = extract_key_information(all_text)
        return key_info
        
    except Exception as e:
        print(f"PDF 처리 중 오류 발생: {e}")
        return None

# OC와 인보이스 데이터 비교 함수
def compare_documents(oc_info, invoice_info):
    """
    OC와 인보이스 정보 비교
    
    Parameters:
    -----------
    oc_info : dict
        OC 문서에서 추출한 정보
    invoice_info : dict
        인보이스 문서에서 추출한 정보
        
    Returns:
    --------
    dict
        비교 결과
    """
    comparison = {
        'matching': True,
        'mismatches': [],
        'items_comparison': []
    }
    
    # 주요 필드 비교
    for field in ['payment_terms', 'incoterms', 'total_price', 'date', 'season', 'brand', 'product_line']:
        if oc_info.get(field) != invoice_info.get(field):
            comparison['matching'] = False
            comparison['mismatches'].append({
                'field': field,
                'oc_value': oc_info.get(field),
                'invoice_value': invoice_info.get(field)
            })
    
    # 제품 항목 비교
    oc_items = {item['style_code']: item for item in oc_info.get('items', [])}
    invoice_items = {item['style_code']: item for item in invoice_info.get('items', [])}
    
    # 모든 스타일 코드 목록
    all_style_codes = set(list(oc_items.keys()) + list(invoice_items.keys()))
    
    for style_code in all_style_codes:
        oc_item = oc_items.get(style_code, {})
        invoice_item = invoice_items.get(style_code, {})
        
        # 해당 스타일이 양쪽 문서에 모두 있는지 확인
        if not oc_item or not invoice_item:
            comparison['matching'] = False
            comparison['items_comparison'].append({
                'style_code': style_code,
                'model': oc_item.get('model') or invoice_item.get('model'),
                'exists_in_oc': bool(oc_item),
                'exists_in_invoice': bool(invoice_item),
                'price_match': False,
                'quantity_match': False
            })
            continue
        
        # 가격 비교
        price_match = oc_item.get('price') == invoice_item.get('price')
        
        # 수량 비교 (전체 수량 기준)
        oc_total_qty = sum(int(qty) for qty in oc_item.get('quantities', []))
        invoice_total_qty = sum(int(qty) for qty in invoice_item.get('quantities', []))
        quantity_match = oc_total_qty == invoice_total_qty
        
        if not price_match or not quantity_match:
            comparison['matching'] = False
        
        comparison['items_comparison'].append({
            'style_code': style_code,
            'model': oc_item.get('model'),
            'oc_price': oc_item.get('price'),
            'invoice_price': invoice_item.get('price'),
            'price_match': price_match,
            'oc_quantity': oc_total_qty,
            'invoice_quantity': invoice_total_qty,
            'quantity_match': quantity_match
        })
    
    return comparison

# 비교 결과를 엑셀 파일로 저장
def save_comparison_to_excel(comparison, output_path):
    """
    비교 결과를 엑셀 파일로 저장
    
    Parameters:
    -----------
    comparison : dict
        비교 결과
    output_path : str
        출력 엑셀 파일 경로
    """
    wb = Workbook()
    
    # 요약 시트
    ws_summary = wb.active
    ws_summary.title = "비교 요약"
    
    # 스타일 설정
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 제목 및 전체 일치 여부
    ws_summary['A1'] = "OC & 인보이스 비교 결과"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "전체 일치 여부:"
    ws_summary['B3'] = "일치" if comparison['matching'] else "불일치"
    ws_summary['B3'].fill = match_fill if comparison['matching'] else mismatch_fill
    
    # 불일치 항목 요약
    if not comparison['matching']:
        ws_summary['A5'] = "불일치 항목 요약"
        ws_summary['A5'].font = Font(bold=True)
        
        headers = ["필드", "OC 값", "인보이스 값", "일치 여부"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=6, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        row = 7
        # 주요 필드 불일치 항목
        for mismatch in comparison['mismatches']:
            ws_summary.cell(row=row, column=1).value = mismatch['field']
            ws_summary.cell(row=row, column=2).value = mismatch['oc_value']
            ws_summary.cell(row=row, column=3).value = mismatch['invoice_value']
            ws_summary.cell(row=row, column=4).value = "불일치"
            
            # 테두리 추가
            for col in range(1, 5):
                ws_summary.cell(row=row, column=col).border = thin_border
            
            # 불일치 표시
            ws_summary.cell(row=row, column=4).fill = mismatch_fill
            
            row += 1
    
    # 제품 비교 시트
    ws_items = wb.create_sheet(title="제품 비교")
    
    # 헤더 설정
    headers = ["스타일 코드", "모델", "OC 가격", "인보이스 가격", "가격 일치", 
               "OC 수량", "인보이스 수량", "수량 일치", "전체 일치"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_items.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = thin_border
    
    # 데이터 입력
    for row, item in enumerate(comparison['items_comparison'], 2):
        # 스타일 코드 및 모델 정보
        ws_items.cell(row=row, column=1).value = item['style_code']
        ws_items.cell(row=row, column=2).value = item['model']
        
        # 가격 비교
        ws_items.cell(row=row, column=3).value = item.get('oc_price', '')
        ws_items.cell(row=row, column=4).value = item.get('invoice_price', '')
        ws_items.cell(row=row, column=5).value = "일치" if item.get('price_match') else "불일치"
        
        if item.get('price_match'):
            ws_items.cell(row=row, column=5).fill = match_fill
        else:
            ws_items.cell(row=row, column=5).fill = mismatch_fill
        
        # 수량 비교
        ws_items.cell(row=row, column=6).value = item.get('oc_quantity', 0)
        ws_items.cell(row=row, column=7).value = item.get('invoice_quantity', 0)
        ws_items.cell(row=row, column=8).value = "일치" if item.get('quantity_match') else "불일치"
        
        if item.get('quantity_match'):
            ws_items.cell(row=row, column=8).fill = match_fill
        else:
            ws_items.cell(row=row, column=8).fill = mismatch_fill
        
        # 전체 일치 여부
        overall_match = item.get('price_match', False) and item.get('quantity_match', False)
        ws_items.cell(row=row, column=9).value = "일치" if overall_match else "불일치"
        
        if overall_match:
            ws_items.cell(row=row, column=9).fill = match_fill
        else:
            ws_items.cell(row=row, column=9).fill = mismatch_fill
        
        # 모든 셀에 테두리 추가
        for col in range(1, 10):
            ws_items.cell(row=row, column=col).border = thin_border
    
    # 열 너비 자동 조정 (수정된 부분: MergedCell 오류 수정)
    for sheet in [ws_summary, ws_items]:
        for column_cells in sheet.columns:
            length = 0
            column_name = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if hasattr(cell, 'value') and cell.value:
                    cell_length = len(str(cell.value))
                    if length < cell_length:
                        length = cell_length
            
            sheet.column_dimensions[column_name].width = length + 2
    
    # 파일 저장
    wb.save(output_path)
    return output_path

# 메인 실행 함수
def compare_oc_and_invoice(oc_path, invoice_path, output_path=None):
    """
    OC와 인보이스 문서를 비교하는 메인 함수
    
    Parameters:
    -----------
    oc_path : str
        OC 문서 파일 경로
    invoice_path : str
        인보이스 문서 파일 경로
    output_path : str, optional
        출력 엑셀 파일 경로
        
    Returns:
    --------
    str
        생성된 엑셀 파일 경로
    """
    if output_path is None:
        output_path = "comparison_enhanced_result.xlsx"
    
    # OC 문서 처리
    print("OC 문서 처리 중...")
    oc_info = extract_info_from_pdf(oc_path, "oc_enhanced_text.txt")
    
    # 인보이스 문서 처리
    print("인보이스 문서 처리 중...")
    invoice_info = extract_info_from_pdf(invoice_path, "invoice_enhanced_text.txt")
    
    # 정보 추출 성공 확인
    if not oc_info or not invoice_info:
        print("문서에서 정보를 추출하지 못했습니다.")
        return None
    
    # 문서 비교
    print("문서 비교 중...")
    comparison = compare_documents(oc_info, invoice_info)
    
    # 결과 저장
    print("비교 결과 저장 중...")
    saved_file = save_comparison_to_excel(comparison, output_path)
    
    print(f"비교 결과가 {saved_file} 파일로 저장되었습니다.")
    
    # 간단한 요약 출력
    print("\n===== 비교 결과 요약 =====")
    print(f"전체 일치 여부: {'일치' if comparison['matching'] else '불일치'}")
    print(f"불일치 항목 수: {len(comparison['mismatches'])}")
    print(f"제품 항목 수: {len(comparison['items_comparison'])}")
    
    return saved_file

# 실행 예시
if __name__ == "__main__":
    oc_path = input("OC 문서 경로 (기본값: oc_document.pdf): ") or "oc_document.pdf"
    invoice_path = input("인보이스 문서 경로 (기본값: invoice_document.pdf): ") or "invoice_document.pdf"
    output_path = input("결과 저장 경로 (기본값: comparison_enhanced_result.xlsx): ") or "comparison_enhanced_result.xlsx"
    
    compare_oc_and_invoice(oc_path, invoice_path, output_path)
    